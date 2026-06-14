"""Local-only bridge between the OneDrive spreadsheet and the repo data files.

NOTE: data/*.json in this repo is the single source of truth. The spreadsheet is
a non-authoritative backup, so `export` (sheet -> repo) will OVERWRITE the
canonical data and should only be run deliberately when the sheet is more current.
Usage:
  python -m scripts.sync export   # xlsx -> data/{fixtures,predictions,players}.json (overwrites canonical data!)
  python -m scripts.sync import   # data/{results,fixtures}.json -> xlsx (Task 10)
"""
import json, sys, os, datetime
import openpyxl
from scripts.score import parse_score

XLSX = os.environ.get("WC_XLSX",
    "/Users/rcairns/Library/CloudStorage/OneDrive-OpenEnergyMarketLimited/RAC/Falcons WC2026 Predictor.xlsx")
DATA = os.path.join(os.path.dirname(__file__), "..", "data")
HEADER_ROW = 3

def _players(ws):
    hdr = [c.value for c in ws[HEADER_ROW]]
    out = []
    for h in hdr[8:]:
        if h is None:
            print(f"WARNING: blank header after {out[-1] if out else 'Res'} — player scan stopped")
            break
        if str(h).startswith("pts "):
            break
        out.append(str(h).strip())
    return out

def read_sheet(ws, stage_col_is_group):
    players = _players(ws)
    fixtures, predictions = [], {}
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if not isinstance(row[0], (int, float)):
            continue
        number = int(row[0])
        kickoff = row[2]
        if isinstance(kickoff, datetime.datetime):
            kickoff = kickoff.date().isoformat()
        elif isinstance(kickoff, datetime.date):
            kickoff = kickoff.isoformat()
        fixtures.append({
            "number": number,
            "stage": "GROUP" if stage_col_is_group else str(row[1] or "").strip(),
            "group": str(row[1] or "").strip() if stage_col_is_group else None,
            "kickoff": kickoff or None,
            "home": str(row[3] or "").strip(), "away": str(row[4] or "").strip(),
            "api_id": None,
        })
        preds = {}
        for i, p in enumerate(players):
            raw = row[8 + i]
            raw = str(raw).strip() if raw is not None else ""
            if not raw:
                continue
            parsed = parse_score(raw)
            if parsed is None:
                print(f"BAD PREDICTION match {number} {p}: {raw!r} (ignored)")
                continue
            preds[p] = f"{parsed[0]}-{parsed[1]}"
        if preds:
            predictions[str(number)] = preds
    return players, fixtures, predictions

def _merge_api_ids(out_dir, fixtures):
    path = os.path.join(out_dir, "fixtures.json")
    if os.path.exists(path):
        with open(path) as fh:
            old = {f["number"]: f for f in json.load(fh)}
        for f in fixtures:
            if f["number"] in old:
                f["api_id"] = old[f["number"]].get("api_id")

def export(xlsx_path, out_dir):
    print("WARNING: data/*.json is the source of truth — this overwrites it from the spreadsheet.")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    players, fixtures, predictions = read_sheet(wb["Group Stage"], True)
    p2, ko_fixtures, ko_preds = read_sheet(wb["Knockouts"], False)
    if set(p2) != set(players):
        sys.exit("ERROR: player sets differ between sheets (someone added/renamed a player in only one sheet):\n"
                 f"  only in Group Stage: {sorted(set(players) - set(p2))}\n"
                 f"  only in Knockouts:   {sorted(set(p2) - set(players))}")
    fixtures += ko_fixtures
    predictions.update(ko_preds)
    _merge_api_ids(out_dir, fixtures)
    with open(os.path.join(out_dir, "players.json"), "w") as fh:
        json.dump(players, fh, indent=1)
    with open(os.path.join(out_dir, "fixtures.json"), "w") as fh:
        json.dump(fixtures, fh, indent=1)
    with open(os.path.join(out_dir, "predictions.json"), "w") as fh:
        json.dump(predictions, fh, indent=1)
    print(f"Exported {len(fixtures)} fixtures, predictions for {len(predictions)} matches, {len(players)} players")
    print("Players:", ", ".join(players))

def _lock_file_present():
    return os.path.exists(os.path.join(os.path.dirname(XLSX), ".~lock." + os.path.basename(XLSX) + "#"))

if __name__ == "__main__":
    if not os.path.exists(XLSX):
        sys.exit(f"Spreadsheet not found at:\n  {XLSX}\n"
                 "It may have been moved — set WC_XLSX to its new path or update the default in scripts/sync.py.")
    if _lock_file_present():
        print("WARNING: spreadsheet appears to be open (lock file present). Close it first.")
    if sys.argv[1:2] == ["export"]:
        export(XLSX, DATA)
    elif sys.argv[1:2] == ["import"]:
        from scripts.sync_import import import_results
        import_results(XLSX, DATA)
    else:
        print(__doc__)
