import json, os
import openpyxl

HEADER_ROW = 3

def _rows_by_number(ws):
    out = {}
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, (int, float)):
            out[int(v)] = r
    return out

def import_results(xlsx_path, data_dir):
    with open(os.path.join(data_dir, "results.json")) as fh:
        results = json.load(fh)
    with open(os.path.join(data_dir, "fixtures.json")) as fh:
        fixtures = {f["number"]: f for f in json.load(fh)}
    wb = openpyxl.load_workbook(xlsx_path)  # keep formulas
    sheets = {"GROUP": wb["Group Stage"], "KO": wb["Knockouts"]}
    rows = {k: _rows_by_number(ws) for k, ws in sheets.items()}
    written = 0
    for mid, res in results.items():
        fx = fixtures.get(int(mid))
        if not fx:
            continue
        key = "GROUP" if fx["stage"] == "GROUP" else "KO"
        row = rows[key].get(int(mid))
        if row and sheets[key].cell(row, 6).value is None:
            sheets[key].cell(row, 6).value = res["home"]
            sheets[key].cell(row, 7).value = res["away"]
            written += 1
    for num, fx in fixtures.items():
        if fx["stage"] == "GROUP" or fx["home"] in ("", "TBD") or fx["away"] in ("", "TBD"):
            continue
        row = rows["KO"].get(num)
        if row:
            sheets["KO"].cell(row, 4).value = fx["home"]
            sheets["KO"].cell(row, 5).value = fx["away"]
            if fx["kickoff"]:
                sheets["KO"].cell(row, 3).value = str(fx["kickoff"])[:10]
    wb.save(xlsx_path)
    print(f"Wrote {written} new results into the spreadsheet")
