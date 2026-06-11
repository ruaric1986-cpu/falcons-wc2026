import openpyxl, json
from scripts.sync import read_sheet, export, _players

def make_wb(path):
    wb = openpyxl.Workbook()
    gs = wb.active; gs.title = "Group Stage"
    gs.append(["title"]); gs.append([""])
    gs.append(["#", "Grp", "Date", "Home Team", "Away Team", "H", "A", "Res",
               "Ruari", "Dave F", "pts Ruari", "pts Dave F"])
    gs.append([1, "A", "2026-06-11", "Mexico", "South Africa", 2, 1, "", "1-0", "2-1", "", ""])
    gs.append([2, "A", "2026-06-12", "Czechia", "South Korea", None, None, "", "0-1", "", "", ""])
    ko = wb.create_sheet("Knockouts")
    ko.append(["title"]); ko.append([""])
    ko.append(["#", "Stage", "Date", "Home Team", "Away Team", "H", "A", "Res",
               "Ruari", "Dave F", "pts Ruari", "pts Dave F"])
    ko.append([73, "R32", "", "TBD", "TBD", None, None, "", "", "", "", ""])
    wb.save(path)

def test_export(tmp_path):
    xlsx = tmp_path / "wc.xlsx"; make_wb(xlsx)
    out = tmp_path / "data"; out.mkdir()
    export(str(xlsx), str(out))
    fixtures = json.loads((out / "fixtures.json").read_text())
    preds = json.loads((out / "predictions.json").read_text())
    players = json.loads((out / "players.json").read_text())
    assert players == ["Ruari", "Dave F"]
    assert fixtures[0] == {"number": 1, "stage": "GROUP", "group": "A",
                           "kickoff": "2026-06-11", "home": "Mexico", "away": "South Africa", "api_id": None}
    assert {f["number"] for f in fixtures} == {1, 2, 73}
    ko = next(f for f in fixtures if f["number"] == 73)
    assert ko["stage"] == "R32" and ko["home"] == "TBD"
    assert preds == {"1": {"Ruari": "1-0", "Dave F": "2-1"}, "2": {"Ruari": "0-1"}}

def test_export_rejects_bad_format(tmp_path, capsys):
    xlsx = tmp_path / "wc.xlsx"; make_wb(xlsx)
    wb = openpyxl.load_workbook(xlsx); wb["Group Stage"]["I5"] = "2:1"; wb.save(xlsx)
    out = tmp_path / "data"; out.mkdir()
    export(str(xlsx), str(out))
    assert "BAD PREDICTION" in capsys.readouterr().out  # reported, not silently dropped
    preds = json.loads((out / "predictions.json").read_text())
    assert "1" in preds          # untouched match still exported
    assert "2" not in preds      # match 2's only prediction was bad -> no entry


def test_blank_player_header_warns(tmp_path, capsys):
    xlsx = tmp_path / "wc.xlsx"; make_wb(xlsx)
    wb = openpyxl.load_workbook(xlsx)
    for sheet in ("Group Stage", "Knockouts"):
        ws = wb[sheet]
        ws.insert_cols(10)  # blank column between the two players on both sheets
    wb.save(xlsx)
    out = tmp_path / "data"; out.mkdir()
    export(str(xlsx), str(out))
    assert "WARNING: blank header" in capsys.readouterr().out
    assert json.loads((out / "players.json").read_text()) == ["Ruari"]
